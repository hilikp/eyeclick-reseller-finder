#!/usr/bin/env python3
"""
EyeClick Reseller Finder — Web App v3.0
Features: Search · Contact Enrichment · Website Links · Email Editor
          Signature · Gmail/Outlook Integration · Sent Tracking · Follow-up Reminders
Run with:  streamlit run app.py
"""

import re, json, time, io, requests, anthropic, os, base64, uuid
import html as html_lib
import urllib.parse
import backend
import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================================================================
# LOGO HELPER  — loads local file as base64; falls back to CDN
# ================================================================
def _logo_img_tag(dark_bg: bool = True) -> str:
    """Return an <img> tag for the EyeClick logo.
    Tries eyeclick_logo.png in the app folder first (works on Cloud too if
    committed); falls back to the official CDN URL."""
    local = os.path.join(os.path.dirname(__file__), "eyeclick_logo.png")
    if os.path.exists(local):
        with open(local, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        src = f"data:image/png;base64,{b64}"
    else:
        src = ("https://cdn.eyeclick.com/logo-light.png" if dark_bg
               else "https://cdn.eyeclick.com/logo-dark.png")
    height = "38px" if dark_bg else "42px"
    return f'<img src="{src}" style="height:{height};object-fit:contain;" alt="EyeClick">'

# ================================================================
# 🔑  API KEYS
# ================================================================
APP_PASSWORD      = st.secrets["APP_PASSWORD"]
ANTHROPIC_API_KEY = st.secrets["ANTHROPIC_API_KEY"]
SERPER_API_KEY    = st.secrets["SERPER_API_KEY"]
HUNTER_API_KEY    = st.secrets.get("HUNTER_API_KEY", "")
GMAIL_USER         = st.secrets.get("GMAIL_USER", "")
GMAIL_APP_PASSWORD = st.secrets.get("GMAIL_APP_PASSWORD", "")
EMAIL_SIGNATURE    = st.secrets.get("EMAIL_SIGNATURE",
    "Best,\n\nYehiel Polatov\nHead of Global Business Development\nBEAM | Obie | Obie for Seniors")
# Email-finder fallback providers (all optional)
APOLLO_API_KEY    = st.secrets.get("APOLLO_API_KEY", "")
SNOV_CLIENT_ID    = st.secrets.get("SNOV_CLIENT_ID", "")
SNOV_CLIENT_SECRET = st.secrets.get("SNOV_CLIENT_SECRET", "")
PROSPEO_API_KEY   = st.secrets.get("PROSPEO_API_KEY", "")

EMAIL_KEYS = {
    "hunter_api_key"    : HUNTER_API_KEY,
    "apollo_api_key"    : APOLLO_API_KEY,
    "snov_client_id"    : SNOV_CLIENT_ID,
    "snov_client_secret": SNOV_CLIENT_SECRET,
    "prospeo_api_key"   : PROSPEO_API_KEY,
}

# ── Import shared constants + pure-logic functions from backend.py ──
from backend import (
    EYECLICK_PROFILE, GOLD_EXAMPLES, QUERY_TEMPLATES, REGIONS, DEFAULT_BLOCKED,
    SENT_LOG_FILE, SEEN_COMPANIES_FILE, FEEDBACK_LOG_FILE, QUEUE_FILE,
    load_sent_log, append_sent_log, mark_followup_done, get_due_followups, already_sent,
    load_feedback_log, save_feedback, is_flagged_wrong_industry,
    load_seen_companies, is_recently_seen, add_to_seen_log,
    is_blocked, validate_website, generate_followup_email,
    load_queue, save_queue, add_to_queue, mark_queue_item,
    send_gmail,
)

# ── Thin wrappers that inject API keys (call sites in app.py stay unchanged) ──
def serper_search(query: str, n: int = 6) -> list:
    return backend.serper_search(query, n, SERPER_API_KEY)

def hunter_search(domain: str) -> dict:
    return backend.hunter_search(domain, HUNTER_API_KEY)

def enrich_contact(client, company: dict) -> dict:
    return backend.enrich_contact(client, company, SERPER_API_KEY, EMAIL_KEYS)

def analyse_companies(client, results, vertical, query, region_label, blocked):
    return backend.analyse_companies(client, results, vertical, query, region_label, blocked)

SENT_LOG_FILE         = "sent_log.json"
SEEN_COMPANIES_FILE   = "seen_companies.json"
FEEDBACK_LOG_FILE     = "feedback_log.json"

# ================================================================
# PAGE SETUP
# ================================================================
st.set_page_config(
    page_title = "EyeClick · Reseller Finder",
    page_icon  = "🎯",
    layout     = "wide",
    initial_sidebar_state = "expanded",
)

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  @import url('https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&display=block');

  html, body, [class*="css"], [data-testid] {
      font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
  }
  #MainMenu, footer, header {visibility: hidden;}
  .block-container {padding-top: 1.5rem !important;}

  /* ── Page background ── */
  [data-testid="stAppViewContainer"] > .main {background: #F5F7F9;}

  /* ── Main header bar ── */
  .ec-header {
      background: linear-gradient(135deg, #101820 0%, #1a1030 100%);
      padding: 1.2rem 1.8rem; border-radius: 14px;
      margin-bottom: 1.6rem;
      display: flex; align-items: center; gap: 1.4rem;
      box-shadow: 0 6px 28px rgba(16,24,32,0.28),
                  0 -2px 0 0 #FF6B9D inset,
                  inset 4px 0 0 0 #CC44DD;
  }
  .ec-header img  {height: 38px; object-fit: contain; flex-shrink: 0;}
  .ec-header-text {line-height: 1.3;}
  .ec-header-text .title {
      font-size: 1.15rem; font-weight: 700; color: #fff; letter-spacing: -.01em;
  }
  .ec-header-text .sub   {font-size: .82rem; color: rgba(255,255,255,.50); margin-top:.15rem;}

  /* ── Result card ── */
  .result-card {
      background: #fff;
      border: 1px solid rgba(16,24,32,0.08);
      border-left: 4px solid #CC44DD;
      border-radius: 12px; padding: 1.1rem 1.4rem;
      margin-bottom: .5rem;
      box-shadow: 0 2px 10px rgba(204,68,221,0.07);
  }

  /* ── Vertical badges ── */
  .badge-seniors       {background:#fce4ec;color:#c2185b;border-radius:20px;padding:2px 10px;font-size:.74rem;font-weight:600;}
  .badge-education     {background:#EEEEF9;color:#5B5CD6;border-radius:20px;padding:2px 10px;font-size:.74rem;font-weight:600;}
  .badge-entertainment {background:#fff3e0;color:#e65100;border-radius:20px;padding:2px 10px;font-size:.74rem;font-weight:600;}
  .badge-healthcare    {background:#e8f5e9;color:#2e7d32;border-radius:20px;padding:2px 10px;font-size:.74rem;font-weight:600;}

  /* ── Score / sent pills ── */
  .score-pill {display:inline-block;background:#5B5CD6;color:#fff;border-radius:20px;padding:2px 11px;font-size:.8rem;font-weight:700;}
  .sent-badge {display:inline-block;background:#1b8a4a;color:#fff;border-radius:20px;padding:2px 10px;font-size:.74rem;font-weight:600;}

  /* ── Follow-up reminder banner ── */
  .reminder-box {
      background:#EEEEF9; border:1px solid #5B5CD6;
      border-left:4px solid #5B5CD6;
      border-radius:10px; padding:.9rem 1.2rem; margin-bottom:1rem;
  }

  /* ── Buttons — pill shape matching eyeclick.com CTAs ── */
  .stButton>button {
      background: #5B5CD6 !important; color: #fff !important;
      border: none !important; border-radius: 200px !important;
      padding: .48rem 1.2rem !important; font-size: .9rem !important;
      font-weight: 600 !important; width: 100% !important;
      letter-spacing: .01em !important;
      transition: background .18s ease, transform .1s ease !important;
  }
  .stButton>button:hover  {background: #4748C4 !important;}
  .stButton>button:active {transform: scale(.97) !important;}

  /* ── Link buttons ── */
  .stLinkButton a {
      background: #5B5CD6 !important; color: #fff !important;
      border-radius: 200px !important; padding: .44rem 1.1rem !important;
      font-weight: 600 !important; font-size: .88rem !important;
      border: none !important; text-decoration: none !important;
      transition: background .18s ease !important;
  }
  .stLinkButton a:hover {background: #4748C4 !important;}

  /* ── Expanders ── */
  div[data-testid="stExpander"] {
      border: 1px solid rgba(16,24,32,0.09) !important;
      border-radius: 10px !important; margin-bottom: .7rem !important;
  }
  div[data-testid="stExpander"] summary {
      font-weight: 500 !important;
  }
  /* Fix: force Material Symbols Rounded font on expander icon spans */
  [data-testid="stExpander"] summary span[class*="icon"],
  [data-testid="stExpander"] summary [data-testid="stExpanderIcon"] {
      font-family: 'Material Symbols Rounded' !important;
      font-feature-settings: 'liga' 1 !important;
      -webkit-font-feature-settings: 'liga' 1 !important;
      font-size: 20px !important;
      line-height: 1 !important;
  }

  /* ── Metric tiles ── */
  [data-testid="metric-container"] {
      background: #fff !important;
      border: 1px solid rgba(16,24,32,0.08) !important;
      border-radius: 10px !important;
  }

  /* ── Sidebar ── */
  [data-testid="stSidebar"] {border-right: 1px solid rgba(16,24,32,0.08);}
  [data-testid="stSidebar"] .stButton>button {border-radius:200px !important;}
</style>
""", unsafe_allow_html=True)

# ================================================================
# PASSWORD GATE
# ================================================================
def login_page():
    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style='text-align:center;margin-bottom:1.8rem;'>
          <div style='background:#101820;display:inline-block;padding:1rem 2rem;
                      border-radius:14px;margin-bottom:1rem;'>
            {_logo_img_tag(dark_bg=True)}
          </div>
          <p style='color:#717171;margin:0;font-size:.95rem;'>Reseller Finder &nbsp;·&nbsp; Team Access</p>
        </div>""", unsafe_allow_html=True)
        pwd = st.text_input("Password", type="password", placeholder="Enter team password")
        if st.button("Sign In", use_container_width=True):
            if pwd == APP_PASSWORD:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password. Please try again.")

if not st.session_state.get("authenticated"):
    login_page()
    st.stop()

# ================================================================
# SIDEBAR
# ================================================================
with st.sidebar:
    st.markdown(f"""
    <div style='background:#101820;border-radius:10px;padding:.7rem 1rem;
                margin-bottom:1.2rem;text-align:center;'>
      {_logo_img_tag(dark_bg=True)}
    </div>""", unsafe_allow_html=True)
    st.markdown("## ⚙️ Settings")

    # Signature
    st.markdown("**📝 Your Email Signature**")
    if "signature" not in st.session_state:
        st.session_state["signature"] = EMAIL_SIGNATURE
    new_sig = st.text_area(
        "sig", value=st.session_state["signature"], height=130,
        placeholder="Best regards,\nYour Name\nEyeClick | Business Development\n+1 234 567 8900",
        label_visibility="collapsed",
    )
    if new_sig != st.session_state["signature"]:
        st.session_state["signature"] = new_sig
        st.success("Signature saved!")

    st.markdown("---")
    st.markdown("**🔄 Deduplication Window**")
    st.caption("Skip companies already found within:")
    dedup_days = st.selectbox(
        "dedup", [7, 14, 30, 60, 90, 0],
        index=2,
        format_func=lambda x: f"{x} days" if x > 0 else "Off (show all)",
        label_visibility="collapsed",
        key="dedup_days",
    )

    # Show seen companies count
    seen_total = len(load_seen_companies())
    if seen_total:
        st.caption(f"📋 {seen_total} companies in history log")

    st.markdown("---")
    st.markdown("**🚫 Blocked Territories**")
    st.caption("Hard-coded: Israel (ALL) · Canada (Seniors)")
    if "extra_blocked" not in st.session_state:
        st.session_state["extra_blocked"] = []
    new_block_country  = st.text_input("Country to block", placeholder="e.g. Germany", key="nb_country")
    new_block_vertical = st.selectbox("Vertical to block", ["ALL","Seniors","Education","Entertainment"], key="nb_vertical")
    if st.button("➕ Add Block", use_container_width=True):
        if new_block_country.strip():
            st.session_state["extra_blocked"].append(
                {"country": new_block_country.strip(), "vertical": new_block_vertical}
            )
            st.success(f"Blocked: {new_block_country.strip()} + {new_block_vertical}")
    if st.session_state["extra_blocked"]:
        for i, b in enumerate(st.session_state["extra_blocked"]):
            c1, c2 = st.columns([3,1])
            c1.caption(f"🚫 {b['country']} · {b['vertical']}")
            if c2.button("✕", key=f"rmblk_{i}"):
                st.session_state["extra_blocked"].pop(i)
                st.rerun()

    st.markdown("---")

    # Sent History
    st.markdown("## 📬 Sent History")
    sent_log = load_sent_log()
    if sent_log:
        for entry in reversed(sent_log[-15:]):
            done = entry.get("follow_up_done", False)
            icon = "✅" if done else "⏰"
            fu   = entry.get("follow_up_date","")
            st.markdown(
                f"**{entry.get('company','')}**  \n"
                f"📅 {entry.get('sent_date','')}  \n"
                f"{icon} Follow-up: {fu}"
            )
            if not done:
                if st.button("Mark follow-up done", key=f"sb_fu_{entry.get('company','')}"):
                    mark_followup_done(entry.get("company",""))
                    st.rerun()
            st.markdown("---")
    else:
        st.info("No emails sent yet.")

    st.markdown("---")
    if st.sidebar.button("🔓 Sign Out"):
        st.session_state["authenticated"] = False
        st.rerun()
    st.markdown("*EyeClick Reseller Finder v3.0*")

# ================================================================
# HEADER
# ================================================================
st.markdown(f"""
<div class="ec-header">
  {_logo_img_tag(dark_bg=True)}
  <div class="ec-header-text">
    <div class="title">Reseller Finder</div>
    <div class="sub">AI-powered worldwide reseller discovery &nbsp;·&nbsp; email outreach &nbsp;·&nbsp; follow-up tracking</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ================================================================
# FOLLOW-UP REMINDERS BANNER
# ================================================================
due_followups = get_due_followups()
if due_followups:
    names = " · ".join(f"**{e.get('company','')}**" for e in due_followups)
    st.markdown(f"""
    <div class="reminder-box">
      ⏰ <strong>{len(due_followups)} follow-up(s) due today</strong> — {names.replace('**','')}
    </div>
    """, unsafe_allow_html=True)
    with st.expander(f"📋 View {len(due_followups)} due follow-up(s)"):
        for entry in due_followups:
            c1, c2, c3 = st.columns([2, 2, 1])
            c1.markdown(f"**{entry.get('company','')}**  \n{entry.get('email','')}")
            c2.markdown(f"First email: {entry.get('sent_date','')}  \nSubject: *{entry.get('subject','')}*")
            if c3.button("✅ Done", key=f"due_{entry.get('company','')}"):
                mark_followup_done(entry.get("company",""))
                st.rerun()
        st.markdown("---")

# ================================================================
# BLOCKED TERRITORIES — get_blocked_territories uses st.session_state so stays here
# ================================================================
def get_blocked_territories() -> list:
    base  = list(DEFAULT_BLOCKED)
    extra = st.session_state.get("extra_blocked", [])
    return base + extra

# ================================================================
# ANTHROPIC CLIENT
# ================================================================
@st.cache_resource
def get_anthropic_client():
    return anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ================================================================
# EXCEL EXPORT
# ================================================================
def build_excel(rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reseller Prospects"
    headers = ["Date","Company","Website","Country","Vertical","Description",
               "Fit Score","Fit Reason","Growth Signals","Contact Name","Title","Email",
               "Email Confidence","LinkedIn","Email Subject","Email Body","Sent?"]
    hfill  = PatternFill(start_color="0057A8", end_color="0057A8", fill_type="solid")
    hfont  = Font(color="FFFFFF", bold=True)
    widths = [11,25,32,14,14,45,9,45,35,20,20,30,14,35,40,80,8]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(1,ci,h); c.fill=hfill; c.font=hfont
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    vcol = {"Seniors":"FCE4EC","Healthcare":"E8F5E9","Education":"E3F2FD","Entertainment":"FFF3E0"}
    sent_companies = {e.get("company") for e in load_sent_log()}
    for r in rows:
        contact = r.get("contact",{})
        ws.append([
            datetime.now().strftime("%Y-%m-%d"),
            r.get("company_name",""), r.get("website",""),
            r.get("country",""), r.get("vertical",""),
            r.get("description",""), r.get("fit_score",""), r.get("fit_reason",""),
            r.get("growth_signals",""),
            contact.get("name",""), contact.get("title",""),
            contact.get("email",""), contact.get("confidence",""), contact.get("linkedin",""),
            r.get("email_subject",""), r.get("email_body",""),
            "Yes" if r.get("company_name") in sent_companies else "No",
        ])
        ri   = ws.max_row
        fill = PatternFill(start_color=vcol.get(r.get("vertical",""),"FFFFFF"),
                           end_color=vcol.get(r.get("vertical",""),"FFFFFF"), fill_type="solid")
        for ci in range(1,len(headers)+1):
            ws.cell(ri,ci).fill = fill
            ws.cell(ri,ci).alignment = Alignment(vertical="top",wrap_text=True)
        ws.row_dimensions[ri].height = 75
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ================================================================
# RESULT CARD v2.0
# ================================================================
def result_card(r: dict, idx: int, key_prefix: str = "all"):
    contact          = r.get("contact", {})
    vertical         = r.get("vertical","")
    badge_cls        = f"badge-{vertical.lower()}"
    score            = r.get("fit_score", 0)
    score_color      = "#2e7d32" if score >= 8 else "#e65100" if score >= 6 else "#999"
    company_name_raw = r.get("company_name","")
    website_raw      = r.get("website","")
    linkedin_raw     = contact.get("linkedin","")
    email_raw        = contact.get("email","")
    sent             = already_sent(company_name_raw)

    def e(v): return html_lib.escape(str(v)) if v else ""

    company_name   = e(company_name_raw)
    country        = e(r.get("country",""))
    description    = e(r.get("description",""))
    fit_reason     = e(r.get("fit_reason",""))
    growth_signals = e(r.get("growth_signals",""))
    contact_name   = e(contact.get("name","—"))
    contact_title  = e(contact.get("title",""))
    confidence     = e(contact.get("confidence",""))

    website_ok     = r.get("website_ok", None)   # None = not checked yet
    li_unverified  = contact.get("linkedin_unverified", False)

    country_html   = f'&nbsp;<span style="color:#666;font-size:.85rem;">🌐 {country}</span>' if country else ""
    conf_html      = f'<span style="color:#888;font-size:.85rem;">{confidence} confidence</span>' if confidence else ""
    sent_html      = '&nbsp;<span class="sent-badge">✅ Email Sent</span>' if sent else ""
    email_link     = (f'<a href="mailto:{e(email_raw)}" style="color:#0057A8;">{e(email_raw)}</a>'
                      if email_raw else "<em style='color:#999;'>not found</em>")
    web_warn       = ('&nbsp;<span style="color:#e65100;font-size:.78rem;font-weight:600;">⚠️ Site unreachable</span>'
                      if website_ok is False else "")
    website_link   = (f'<a href="{e(website_raw)}" target="_blank" style="color:#0057A8;font-size:.88rem;">'
                      f'🌐 {e(website_raw)}</a>{web_warn}' if website_raw else "")
    has_growth     = growth_signals and growth_signals.lower() not in ("none detected","none","")
    growth_html    = (f'<div style="margin:.3rem 0;background:#fff8e1;border-left:3px solid #ffb300;'
                      f'padding:.3rem .6rem;border-radius:4px;font-size:.82rem;color:#7a5800;">'
                      f'📈 <strong>Growth signal:</strong> {growth_signals}</div>'
                      if has_growth else "")

    # Evidence snippets
    snippets       = r.get("evidence_snippets", [])
    snippets_html  = ""
    if snippets:
        items = "".join(f'<li style="margin:.15rem 0;">{e(s)}</li>' for s in snippets[:2])
        snippets_html = (f'<div style="margin:.3rem 0;background:#f0f4fa;border-left:3px solid #0057A8;'
                         f'padding:.3rem .6rem;border-radius:4px;font-size:.82rem;color:#333;">'
                         f'🔍 <strong>Evidence:</strong><ul style="margin:.2rem 0;padding-left:1.2rem;">{items}</ul></div>')

    # ── Main card ──
    st.markdown(f"""
    <div class="result-card">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:.4rem;">
        <div>
          <strong style="font-size:1.08rem;">{idx}. {company_name}</strong>
          &nbsp;<span class="{badge_cls}">{vertical}</span>
          {country_html}{sent_html}
        </div>
        <span class="score-pill" style="background:{score_color};">{score}/10</span>
      </div>
      {('<div style="margin:.25rem 0 .3rem;">' + website_link + '</div>') if website_link else ''}
      <p style="margin:.4rem 0 .25rem;color:#444;font-size:.92rem;">{description}</p>
      <p style="margin:0 0 .25rem;color:#555;font-size:.85rem;"><em>{fit_reason}</em></p>
      {snippets_html}
      {growth_html}
      <hr style="margin:.5rem 0;border-color:#eee;">
      <div style="display:flex;gap:1.4rem;flex-wrap:wrap;font-size:.9rem;align-items:center;">
        <span>👤 <strong>{contact_name}</strong>{(" · " + contact_title) if contact_title else ""}</span>
        <span>📧 {email_link}</span>
        {conf_html}
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── LinkedIn + action row ──
    has_linkedin = linkedin_raw and linkedin_raw.startswith("http")
    if has_linkedin:
        unverified_note = " ⚠️ *profile match unverified — please confirm*" if li_unverified else ""
        st.markdown(f"&nbsp;&nbsp;&nbsp;[🔗 LinkedIn Profile]({linkedin_raw}){unverified_note}",
                    unsafe_allow_html=True)

    # ── Email editor expander ──
    email_label = "✉️  Edit & Send Email" + (" — ✅ Sent" if sent else "")
    with st.expander(f"{email_label} · {company_name_raw}"):

        # Subject
        subj_key  = f"subj_{key_prefix}_{hash(website_raw)}"
        if subj_key not in st.session_state:
            st.session_state[subj_key] = r.get("email_subject","")
        subject = st.text_input("Subject line", key=subj_key)

        # Body — draft only, NO signature baked in (signature added live below)
        body_key = f"body_{key_prefix}_{hash(website_raw)}"
        if body_key not in st.session_state:
            st.session_state[body_key] = r.get("email_body","")
        body = st.text_area("Email body (fully editable)", key=body_key, height=220)

        # Signature — always reflects current sidebar value, never baked into body
        sig = st.session_state.get("signature","")
        if sig:
            st.markdown("**✍️ Signature** *(edit in Settings panel)*")
            st.code(sig, language=None)
        else:
            st.caption("💡 Add your signature in the ⚙️ **Settings** panel on the left — it will appear here automatically.")

        # Full email = greeting + body + signature (combined only for sending)
        contact_name_raw = contact.get("name","") if contact else ""
        full_email = _greeting(contact_name_raw) + "\n\n" + body + ("\n\n" + sig if sig else "")

        # ── 4 action buttons ──
        a1, a2, a3, a4 = st.columns(4)

        # Open in email client (mailto)
        if email_raw:
            mailto = ("mailto:" + email_raw
                      + "?subject=" + urllib.parse.quote(subject)
                      + "&body="    + urllib.parse.quote(full_email))
            a1.link_button("📨 Open in Email Client", mailto, use_container_width=True)
        else:
            a1.info("No email found")

        # Mark as sent + set follow-up
        if not sent:
            if a2.button("✅ Mark as Sent + Reminder", key=f"mark_{key_prefix}_{hash(website_raw)}", use_container_width=True):
                fu_date = (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d")
                append_sent_log({
                    "company"        : company_name_raw,
                    "website"        : website_raw,
                    "email"          : email_raw,
                    "subject"        : subject,
                    "sent_date"      : datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "follow_up_date" : fu_date,
                    "follow_up_done" : False,
                })
                st.success(f"✅ Logged! Follow-up reminder set for **{fu_date}**")
                st.rerun()
        else:
            a2.success("✅ Already sent")

        # Visit LinkedIn
        if has_linkedin:
            a3.link_button("👁 Visit LinkedIn", linkedin_raw, use_container_width=True)

        # Add to outreach queue
        if email_raw:
            if a4.button("➕ Queue", key=f"queue_{key_prefix}_{hash(website_raw)}", use_container_width=True):
                item = {
                    "id"           : str(uuid.uuid4()),
                    "type"         : "initial",
                    "company_name" : company_name_raw,
                    "website"      : website_raw,
                    "vertical"     : vertical,
                    "contact_name" : contact.get("name",""),
                    "contact_email": email_raw,
                    "subject"      : st.session_state.get(subj_key, r.get("email_subject","")),
                    "body"         : st.session_state.get(body_key, r.get("email_body","")),
                    "queued_date"  : datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "status"       : "pending",
                    "sent_date"    : None,
                }
                st.toast("✅ Added to Outreach Queue" if add_to_queue(item) else "⚠️ Already in queue")

        # ── Copy full email preview ──
        full_preview = body + ("\n\n" + sig if sig else "")
        with st.expander("📋 Preview & copy full email"):
            st.code(full_preview, language=None)

        # Follow-up email (if already sent)
        if sent:
            st.markdown("---")
            st.markdown("**🔄 Send Follow-up Email**")
            fu_subj_key = f"fu_subj_{key_prefix}_{hash(website_raw)}"
            fu_body_key = f"fu_body_{key_prefix}_{hash(website_raw)}"
            if fu_subj_key not in st.session_state:
                st.session_state[fu_subj_key] = f"Following up: {subject}"
            if fu_body_key not in st.session_state:
                first = contact.get("name","").split()[0] if contact.get("name") else "there"
                st.session_state[fu_body_key] = (
                    f"Hi {first},\n\n"
                    f"I wanted to follow up on my previous message regarding EyeClick's interactive "
                    f"projection systems. Did you get a chance to review it?\n\n"
                    f"I'd love to schedule a quick 15-minute call to explore if there's a potential "
                    f"fit for a reseller partnership.\n\n"
                    f"Looking forward to hearing from you."
                )
            fu_subj = st.text_input("Follow-up subject", key=fu_subj_key)
            fu_body = st.text_area("Follow-up body (no signature — added automatically below)", key=fu_body_key, height=180)

            # AI generate button
            ai_gen_key = f"ai_gen_{key_prefix}_{hash(website_raw)}"
            if st.button("🤖 Generate AI Follow-up", key=ai_gen_key, use_container_width=True):
                with st.spinner("Writing personalised follow-up…"):
                    generated = generate_followup_email(
                        get_anthropic_client(), r, contact, subject
                    )
                if generated:
                    st.session_state[fu_body_key] = generated
                    st.rerun()
                else:
                    st.warning("Could not generate — try again.")

            # Signature (live, never baked into body)
            sig = st.session_state.get("signature","")
            fu_full = fu_body + ("\n\n" + sig if sig else "")

            # Full follow-up preview
            with st.expander("📋 Preview & copy follow-up email"):
                st.code(fu_full, language=None)

            if email_raw:
                fu_mailto = ("mailto:" + email_raw
                             + "?subject=" + urllib.parse.quote(fu_subj)
                             + "&body="    + urllib.parse.quote(fu_full))
                b1, b2 = st.columns(2)
                b1.link_button("📨 Open Follow-up in Email Client", fu_mailto, use_container_width=True)
                if b2.button("✅ Mark Follow-up Sent", key=f"fu_done_{key_prefix}_{hash(website_raw)}", use_container_width=True):
                    mark_followup_done(company_name_raw)
                    st.success("Follow-up marked as done!")
                    st.rerun()

    # ── Report Issue ──
    report_options = [
        "a. Incorrect details — website or contact info is wrong",
        "b. LinkedIn profile is not correct / couldn't be verified",
        "c. Wrong industry — this company is irrelevant (remove permanently)",
    ]
    with st.expander(f"🚩 Report an Issue · {company_name_raw}"):
        st.markdown("Help improve search quality by flagging a problem with this result:")
        chosen = st.radio("Issue type:", report_options,
                          key=f"report_radio_{key_prefix}_{hash(website_raw)}",
                          label_visibility="collapsed")
        if st.button("🚩 Submit Report", key=f"report_btn_{key_prefix}_{hash(website_raw)}",
                     use_container_width=True):
            code = "details" if chosen.startswith("a") else "linkedin" if chosen.startswith("b") else "industry"
            save_feedback(company_name_raw, website_raw, code)
            if code == "industry":
                current = st.session_state.get("last_results", [])
                st.session_state["last_results"] = [
                    c for c in current if c.get("website","") != website_raw
                ]
                st.success("🗑️ Removed from results. This company will be skipped in future searches.")
                st.rerun()
            elif code == "linkedin":
                st.success("✅ LinkedIn issue noted. Please use the correct profile if you find it.")
            else:
                st.success("✅ Details issue reported — thank you!")

# ================================================================
# OUTREACH QUEUE TAB RENDERER
def _greeting(contact_name: str) -> str:
    first = contact_name.strip().split()[0] if contact_name.strip() else ""
    return f"Hi {first}," if first else "Hi,"

# ================================================================
def _render_outreach_queue_tab():
    queue      = load_queue()
    pending    = [i for i in queue if i["status"] == "pending"]
    today_str  = datetime.now().strftime("%Y-%m-%d")
    sent_today = [i for i in queue if i["status"] == "sent"
                  and (i.get("sent_date") or "").startswith(today_str)]
    skipped    = [i for i in queue if i["status"] == "skipped"]

    cb1, cb2, cb3 = st.columns(3)
    cb1.metric("⏳ Pending",    len(pending))
    cb2.metric("✅ Sent Today", len(sent_today))
    cb3.metric("⏭ Skipped",    len(skipped))

    if not pending:
        st.info("No pending emails. Use the ➕ Queue button on any result card to stage emails for sending.")
        return

    sig = st.session_state.get("signature", "")

    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        st.warning("⚠️ Gmail not configured. Add **GMAIL_USER** and **GMAIL_APP_PASSWORD** to your secrets to enable sending.")

    if st.button("🚀 Send All Pending", use_container_width=True, key="send_all_pending"):
        if not GMAIL_USER or not GMAIL_APP_PASSWORD:
            st.error("Set Gmail credentials first (see Settings).")
        else:
            sent_count, errors = 0, []
            for item in pending:
                full_body = _greeting(item.get("contact_name","")) + "\n\n" + item["body"]
                ok = send_gmail(
                    to=item["contact_email"], subject=item["subject"],
                    body=full_body, signature=sig,
                    gmail_user=GMAIL_USER, gmail_app_password=GMAIL_APP_PASSWORD,
                )
                if ok:
                    mark_queue_item(item["id"], "sent")
                    fu_date = (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d")
                    append_sent_log({
                        "company"       : item["company_name"],
                        "website"       : item["website"],
                        "email"         : item["contact_email"],
                        "subject"       : item["subject"],
                        "sent_date"     : datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "follow_up_date": fu_date,
                        "follow_up_done": False,
                    })
                    if item["type"] == "followup":
                        mark_followup_done(item["company_name"])
                    sent_count += 1
                else:
                    errors.append(item["company_name"])
            if sent_count:
                st.success(f"✅ {sent_count} email(s) sent!")
            if errors:
                st.error(f"Failed to send: {', '.join(errors)}")
            st.rerun()

    st.markdown("---")

    for item in pending:
        has_email  = bool(item.get("contact_email"))
        email_disp = item["contact_email"] if has_email else "⚠️ no email found"
        with st.expander(f"{'✉️' if has_email else '🔍'} **{item['company_name']}** — {email_disp}  ·  {item.get('vertical','')}"):
            if not has_email:
                st.warning("No email found for this company. You can:\n- Add an email manually below\n- Visit their website or LinkedIn to find a contact\n- Skip this entry")
                manual_email = st.text_input("Add email manually", placeholder="contact@company.com",
                                             key=f"manual_email_{item['id']}")
                if manual_email and st.button("💾 Save email", key=f"save_email_{item['id']}"):
                    q_all = load_queue()
                    for qi in q_all:
                        if qi["id"] == item["id"]:
                            qi["contact_email"] = manual_email
                    save_queue(q_all)
                    st.success("Email saved!")
                    st.rerun()
                if item.get("website"):
                    st.markdown(f"🌐 [Visit website]({item['website']})")

            qsubj_key = f"q_subj_{item['id']}"
            qbody_key = f"q_body_{item['id']}"
            if qsubj_key not in st.session_state:
                st.session_state[qsubj_key] = item["subject"]
            if qbody_key not in st.session_state:
                st.session_state[qbody_key] = item["body"]

            edited_subj = st.text_input("Subject", key=qsubj_key)
            edited_body = st.text_area("Body", key=qbody_key, height=180)

            # Persist edits back to queue file immediately
            if edited_subj != item["subject"] or edited_body != item["body"]:
                q_all = load_queue()
                for qi in q_all:
                    if qi["id"] == item["id"]:
                        qi["subject"] = edited_subj
                        qi["body"]    = edited_body
                save_queue(q_all)

            if sig:
                st.caption("✍️ *Signature will be appended at send time*")

            c1, c2 = st.columns(2)
            send_disabled = not has_email
            if c1.button("📤 Send Now", key=f"send_now_{item['id']}", use_container_width=True,
                         disabled=send_disabled):
                if not GMAIL_USER or not GMAIL_APP_PASSWORD:
                    st.error("Set Gmail credentials first.")
                else:
                    full_body = _greeting(item.get("contact_name","")) + "\n\n" + edited_body
                    ok = send_gmail(
                        to=item["contact_email"], subject=edited_subj,
                        body=full_body, signature=sig,
                        gmail_user=GMAIL_USER, gmail_app_password=GMAIL_APP_PASSWORD,
                    )
                    if ok:
                        mark_queue_item(item["id"], "sent")
                        fu_date = (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d")
                        append_sent_log({
                            "company"       : item["company_name"],
                            "website"       : item["website"],
                            "email"         : item["contact_email"],
                            "subject"       : edited_subj,
                            "sent_date"     : datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "follow_up_date": fu_date,
                            "follow_up_done": False,
                        })
                        if item["type"] == "followup":
                            mark_followup_done(item["company_name"])
                        st.success(f"✅ Sent to {item['contact_email']}")
                        st.rerun()
                    else:
                        st.error("Send failed — check Gmail credentials in secrets.")
            if c2.button("⏭ Skip", key=f"skip_{item['id']}", use_container_width=True):
                mark_queue_item(item["id"], "skipped")
                st.rerun()

# ================================================================
# SEARCH FORM
# ================================================================
with st.container():
    col1, col2, col3, col4 = st.columns([2, 2, 1.2, 1])
    with col1:
        st.markdown("**Vertical (market segment)**")
        sel_h = st.checkbox("👴 Seniors",        value=True)
        sel_e = st.checkbox("🏫 Education",     value=True)
        sel_n = st.checkbox("🎯 Entertainment", value=False)
    with col2:
        st.markdown("**Region**")
        region_label = st.selectbox("", list(REGIONS.keys()), label_visibility="collapsed")
        region_kw    = REGIONS[region_label]
    with col3:
        st.markdown("**Number of results**")
        num_results = st.slider("", min_value=5, max_value=30, value=10, step=5,
                                label_visibility="collapsed")
    with col4:
        st.markdown("<br>", unsafe_allow_html=True)
        search_clicked = st.button("🔍  SEARCH", use_container_width=True)

selected_verticals = (["Seniors"]       if sel_h else []) + \
                     (["Education"]     if sel_e else []) + \
                     (["Entertainment"] if sel_n else [])

st.markdown("---")

# ================================================================
# SEARCH LOGIC
# ================================================================
if search_clicked:
    if not selected_verticals:
        st.warning("Please select at least one vertical.")
        st.stop()

    import itertools
    client       = get_anthropic_client()
    all_companies: list = []
    seen_sites  : set  = set()
    today        = datetime.now().strftime("%Y-%m-%d")
    status_box   = st.empty()
    progress_bar = st.progress(0)
    log_box      = st.empty()
    log_lines: list = []

    v_cycle = itertools.cycle(selected_verticals)
    q_pool  = {v: [q.format(region=region_kw).strip() for q in QUERY_TEMPLATES[v]]
               for v in selected_verticals}
    blocked = get_blocked_territories()
    attempt = 0

    while len(all_companies) < num_results and attempt < 40:
        v = next(v_cycle)
        if not q_pool[v]:
            q_pool[v] = [q.format(region=region_kw).strip() for q in QUERY_TEMPLATES[v]]
        query = q_pool[v].pop(0)
        attempt += 1
        pct = min(int((len(all_companies) / num_results) * 70), 70)
        progress_bar.progress(pct)
        status_box.info(f"🔍 Searching ({v}): *{query}* — {len(all_companies)}/{num_results} found…")
        results   = serper_search(query, 6)
        if not results:
            continue
        companies = analyse_companies(client, results, v, query, region_label, blocked)
        dedup     = st.session_state.get("dedup_days", 30)
        new       = [c for c in companies
                     if c.get("website","") not in seen_sites
                     and not is_recently_seen(c.get("website",""), dedup)
                     and not is_blocked(c.get("country",""), v, blocked)
                     and not is_flagged_wrong_industry(c.get("website",""))]
        for c in new:
            seen_sites.add(c.get("website",""))
        all_companies.extend(new)
        log_lines.append(f"✅ {v} · {len(new)} co. — `{query}`")
        log_box.markdown("\n".join(log_lines[-6:]))
        time.sleep(0.8)

    final = all_companies[:num_results]
    status_box.info(f"📇 Enriching contacts for {len(final)} companies…")
    for i, company in enumerate(final):
        progress_bar.progress(70 + int((i / max(len(final),1)) * 28))
        status_box.info(f"📇 Finding contact: **{company.get('company_name','')}** ({i+1}/{len(final)})…")
        company["contact"]    = enrich_contact(client, company)
        company["website_ok"] = validate_website(company.get("website",""))
        time.sleep(0.6)

    progress_bar.progress(100)
    status_box.success(f"✅ Done! **{len(final)} reseller candidates** found with contact details.")
    log_box.empty()

    # Save to seen-companies log so future searches skip these
    add_to_seen_log(final)

    # Store results in session so they survive reruns (e.g. mark-as-sent buttons)
    st.session_state["last_results"] = final
    st.session_state["last_date"]    = today
    st.session_state["last_region"]  = region_label

# ================================================================
# DISPLAY RESULTS (from session state — survives reruns)
# ================================================================
final = st.session_state.get("last_results")
if final:
    today        = st.session_state.get("last_date", datetime.now().strftime("%Y-%m-%d"))
    region_label = st.session_state.get("last_region","")

    st.markdown(f"## Results · {region_label.strip()} · {today}")

    # Stats bar
    groups = {"Seniors":[], "Education":[], "Entertainment":[]}
    for r in final:
        groups.get(r.get("vertical",""), []).append(r)
    avg_score = round(sum(r.get("fit_score",0) for r in final) / max(len(final),1), 1)
    with_email = sum(1 for r in final if r.get("contact",{}).get("email"))
    s1, s2, s3, s4, s5, s6 = st.columns(6)
    s1.metric("Total Found", len(final))
    s2.metric("👴 Seniors",       len(groups["Seniors"]))
    s3.metric("🏫 Education",     len(groups["Education"]))
    s4.metric("🎯 Entertainment", len(groups["Entertainment"]))
    s5.metric("Avg Fit Score",    f"{avg_score}/10")
    s6.metric("With Email",       with_email)
    st.markdown("---")

    tabs   = st.tabs(["📋 All Results", "👴 Seniors", "🏫 Education", "🎯 Entertainment", "📬 Outreach Queue"])

    with tabs[0]:
        for i, r in enumerate(final, 1):
            result_card(r, i, key_prefix="all")

    for tab, vertical in zip(tabs[1:4], ["Seniors","Education","Entertainment"]):
        with tab:
            vlist = groups[vertical]
            if vlist:
                for i, r in enumerate(vlist, 1):
                    result_card(r, i, key_prefix=vertical.lower())
            else:
                st.info(f"No {vertical} companies found in this search.")

    with tabs[4]:
        _render_outreach_queue_tab()

    st.markdown("---")
    excel_bytes = build_excel(final)
    st.download_button(
        label     = "⬇️  Download Full Excel Report",
        data      = excel_bytes,
        file_name = f"eyeclick_resellers_{today}.xlsx",
        mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width = True,
    )
else:
    st.markdown("""
    <div style='text-align:center;padding:2rem;color:#888;'>
      <span style='font-size:3rem;'>🔍</span>
      <p style='font-size:1.1rem;margin-top:1rem;'>
        Select your verticals and region above, then hit <strong>SEARCH</strong><br>
        to discover ideal EyeClick reseller partners worldwide.
      </p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("### 📬 Outreach Queue")
    _render_outreach_queue_tab()
